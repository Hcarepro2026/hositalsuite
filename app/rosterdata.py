"""Unified roster engine — one roster for the whole hospital.

WHY THIS EXISTS
---------------
The suite used to carry TWO rosters on TWO pages that could not see each other:

  * "Duty Roster"  (/roster)      — one Admin Manager per day, hospital-wide.
  * "Dept Roster"  (/dept-roster) — a department, a date, a shift, and exactly
                                    two staff columns (`staff1`, `staff2`).

That split caused four real problems for a hospital:

  1. A ward with nine nurses on nights could not be rostered — there were only
     two staff slots.
  2. Sections and units (A&E › Triage) could not own a roster at all, even
     though the structure page lets you create them.
  3. Administrative departments (Procurement, Internal Audit, Finance &
     Accounts, ICT, Admin/HR) do not run shifts. Forcing them into DAY/NIGHT
     produced a roster nobody believed.
  4. Leave was invisible. Nothing stopped an admin from rostering a nurse who
     was on annual leave, because the roster had no idea leave existed.

This module is the shared engine behind the single "Roster" page: date ranges,
scope resolution, shift rules, leave, file parsing, validation-before-write and
the commit step. The view layer (app/views/roster.py) only handles HTTP.

DESIGN RULES
------------
* NOTHING IS WRITTEN UNTIL A HUMAN HAS SEEN IT. Every upload produces a preview
  with a per-row verdict and a plain-English reason for each rejection.
* PREVIEWS ARE NOT KEPT IN THE SESSION COOKIE. A Flask session holds ~4KB —
  about 24 rows. A month of ward roster is thousands. Previews live in durable
  storage under a token, exactly like the staff bulk upload.
* THE ADMIN-MANAGER ROSTER STAYS IN `duty_roster`. Duty reminders, overdue
  inspection chasing and the compliance report all read that table. Moving it
  would have been a rewrite with nothing to show for it, so the unified page
  reads and writes it in place and simply displays it beside everything else.
"""
from __future__ import annotations

import csv
import io
import json
import re
import secrets
from datetime import date, datetime, timedelta

from .models import (ALL_SHIFT_CODES, DEPT_SHIFTS, LEAVE_CODES, LEAVE_LABELS,
                     Department, DeptRosterEntry, DutyRoster, RosterEntry,
                     Section, Unit, User, db, now_naive)

MAX_ROWS = 5000          # a month of a big ward, with headroom
MAX_RANGE_DAYS = 366     # refuse absurd date ranges rather than build them

DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y",
                "%d %b %Y", "%d %B %Y", "%b %d %Y", "%d.%m.%Y")


# ------------------------------------------------------------------ dates
def parse_date(raw) -> date | None:
    """Accept the date formats Nigerian hospital spreadsheets actually contain."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    raw = str(raw).strip()
    if not raw:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


# The founder asked to be able to pick "a day, 7 days, 2 weeks, 3 weeks, a month".
RANGE_PRESETS = [
    ("today", "Today", 1),
    ("7", "Next 7 days", 7),
    ("14", "Next 2 weeks", 14),
    ("21", "Next 3 weeks", 21),
    ("30", "Next 30 days", 30),
    ("month", "This calendar month", 0),
    ("custom", "Choose my own dates", 0),
]
PRESET_DAYS = {k: n for k, _, n in RANGE_PRESETS if n}


def resolve_range(preset: str, raw_from: str = "", raw_to: str = "",
                  today: date | None = None) -> tuple[date, date, str]:
    """Turn a preset (or two typed dates) into a real start/end pair.

    Always returns a sane range: start <= end, never longer than a year. A bad
    or missing input falls back to the next 7 days rather than erroring — the
    roster page must always render something.
    """
    today = today or now_naive().date()
    preset = (preset or "7").strip()

    if preset == "custom":
        start = parse_date(raw_from) or today
        end = parse_date(raw_to) or (start + timedelta(days=6))
    elif preset == "month":
        start = today.replace(day=1)
        end = (start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    elif preset == "today":
        start = end = today
    else:
        days = PRESET_DAYS.get(preset, 7)
        start, end = today, today + timedelta(days=days - 1)

    if end < start:
        start, end = end, start
    if (end - start).days > MAX_RANGE_DAYS:
        end = start + timedelta(days=MAX_RANGE_DAYS)
    label = (start.strftime("%a %d %b %Y") if start == end
             else f"{start.strftime('%a %d %b %Y')} → {end.strftime('%a %d %b %Y')}")
    return start, end, label


def days_between(start: date, end: date):
    d = start
    while d <= end:
        yield d
        d += timedelta(days=1)


# ------------------------------------------------------------------ shifts
def shifts_for(mode: str | None) -> list[tuple[str, str]]:
    return DEPT_SHIFTS.get(mode or "two_12h", DEPT_SHIFTS["two_12h"])


def shift_codes(mode: str | None) -> list[str]:
    return [s[0] for s in shifts_for(mode)]


def shift_hours(code: str) -> str:
    for shifts in DEPT_SHIFTS.values():
        for c, hours in shifts:
            if c == code:
                return hours
    return ""


WEEKEND = (5, 6)  # Saturday, Sunday


# The hospital-wide Admin Manager roster has no shifts and no weekends off — the
# hospital runs every day. It is deliberately NOT a working pattern you can pick
# for a department; it is what the ORG scope always uses.
ORG_SHIFT = "ADMIN"


def office_mode_violation(mode: str | None, day: date) -> str | None:
    """Office departments do not work weekends — say so instead of saving it.

    Never applies to the Admin Manager roster: somebody is on duty on a Sunday.
    """
    if (mode or "") == "office" and day.weekday() in WEEKEND:
        return (f"{day.strftime('%A %d %b')} is a weekend and this department is set to "
                "'Office hours, Monday to Friday'. Change the department's roster pattern "
                "first if it really does work weekends.")
    return None


# ------------------------------------------------------------------ scope
def resolve_scope(org_id: int, scope: str, department_id=None, section_id=None,
                  unit_id=None) -> tuple[dict, list[str]]:
    """Validate the place a roster line belongs to. Returns (place, errors).

    Tenant isolation is enforced here: a section or unit from another hospital
    is rejected outright, and a section must genuinely sit inside the chosen
    department (otherwise the roster would claim Pharmacy owns an A&E unit).
    """
    errors: list[str] = []
    scope = (scope or "DEPARTMENT").upper()
    if scope not in ("ORG", "DEPARTMENT", "SECTION", "UNIT"):
        return {}, ["Choose who owns this roster: Admin Manager, department, section or unit."]

    if scope == "ORG":
        return {"scope": "ORG", "department_id": None, "section_id": None, "unit_id": None}, []

    dept = db.session.get(Department, department_id) if department_id else None
    if not dept or dept.org_id != org_id:
        return {}, ["Choose a department."]

    section = unit = None
    if scope in ("SECTION", "UNIT"):
        section = db.session.get(Section, section_id) if section_id else None
        if not section or section.org_id != org_id:
            errors.append("Choose a section inside that department.")
        elif section.department_id != dept.id:
            errors.append(f"'{section.name}' is not a section of {dept.name}.")
    if scope == "UNIT":
        unit = db.session.get(Unit, unit_id) if unit_id else None
        if not unit or unit.org_id != org_id:
            errors.append("Choose a unit inside that section.")
        elif section and unit.section_id != section.id:
            errors.append(f"'{unit.name}' is not a unit of {section.name}.")

    if errors:
        return {}, errors
    return {"scope": scope, "department_id": dept.id,
            "section_id": section.id if section else None,
            "unit_id": unit.id if unit else None}, []


def can_manage(user, place: dict) -> bool:
    """Super Admin manages everything; an HOD manages only their department."""
    if user.is_super:
        return True
    if getattr(user, "role", "") == "HOD" and place.get("department_id"):
        dept = db.session.get(Department, place["department_id"])
        return bool(dept and dept.hod_user_id == user.id)
    return False


def visible_departments(user):
    q = db.session.query(Department).filter_by(org_id=user.org_id, active=True)
    if getattr(user, "role", "") == "HOD" and not user.is_super:
        q = q.filter(Department.hod_user_id == user.id)
    return q.order_by(Department.name).all()


# ------------------------------------------------------------------ people
def staff_index(org_id: int) -> dict[str, User]:
    """Look people up the way a spreadsheet spells them.

    Real files contain "MRS ABATAN L.F", "CNO Ogunleye" and "pharm kareem".
    We index the exact name, the name without its title, and the username, so a
    match does not depend on the typist.
    """
    idx: dict[str, User] = {}
    for u in db.session.query(User).filter_by(org_id=org_id, active=True).all():
        for key in {_norm_person(u.name), _norm_person(_strip_title(u.name)),
                    (u.username or "").strip().lower()}:
            if key and key not in idx:
                idx[key] = u
    return idx


_TITLES = ("dr", "mr", "mrs", "miss", "ms", "prof", "pharm", "engr", "cno", "adns",
           "dns", "sr", "matron", "nurse", "rev", "alhaji", "alhaja", "chief")


def _norm_person(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (name or "").lower()).strip()


def _strip_title(name: str) -> str:
    parts = _norm_person(name).split()
    while parts and parts[0] in _TITLES:
        parts = parts[1:]
    return " ".join(parts)


def find_person(idx: dict, raw: str) -> User | None:
    for key in (_norm_person(raw), _norm_person(_strip_title(raw))):
        if key and key in idx:
            return idx[key]
    return None


# ------------------------------------------------------------------ leave
def leave_on(org_id: int, user_id: int, day: date) -> RosterEntry | None:
    return (db.session.query(RosterEntry)
            .filter_by(org_id=org_id, user_id=user_id, duty_date=day, kind="LEAVE")
            .first())


def normalise_leave(raw: str) -> str | None:
    """Accept 'Annual', 'annual leave', 'AL', 'sick', 'off' ..."""
    t = re.sub(r"[^a-z]+", "", (raw or "").lower()).replace("leave", "")
    aliases = {"al": "ANNUAL", "annual": "ANNUAL", "casual": "CASUAL", "cl": "CASUAL",
               "sick": "SICK", "sl": "SICK", "study": "STUDY", "maternity": "MATERNITY",
               "mat": "MATERNITY", "compassionate": "COMPASSIONATE",
               "exam": "EXAM", "examination": "EXAM", "off": "OFF", "offduty": "OFF"}
    if t in aliases:
        return aliases[t]
    up = (raw or "").strip().upper()
    return up if up in LEAVE_CODES else None


# ------------------------------------------------------------------ file parsing
_COLS = {
    "name": ["name", "names", "staff", "staff name", "full name", "officer",
             "employee name", "admin manager", "person"],
    "date": ["date", "duty date", "day date", "from", "start", "start date",
             "date from", "duty_date"],
    "end": ["end", "to", "end date", "date to", "until"],
    "shift": ["shift", "duty", "tour", "shift/leave", "type"],
    "leave": ["leave", "leave type", "leave_type", "absence"],
    "department": ["department", "dept"],
    "section": ["section"],
    "unit": ["unit"],
    "note": ["note", "notes", "remark", "remarks", "comment", "duty assignment"],
}


def _header_map(headers: list[str]) -> dict[str, int]:
    norm = [re.sub(r"\s+", " ", (h or "").strip().lower()) for h in headers]
    out: dict[str, int] = {}
    for field, aliases in _COLS.items():
        for i, h in enumerate(norm):
            if h in aliases and field not in out:
                out[field] = i
    return out


def parse_file(file_storage) -> tuple[list[dict], str | None]:
    """Read CSV or XLSX into dicts. Returns (rows, error_message)."""
    filename = (getattr(file_storage, "filename", "") or "").lower()
    try:
        if filename.endswith(".csv"):
            text = file_storage.read().decode("utf-8-sig", errors="replace")
            table = list(csv.reader(io.StringIO(text)))
        elif filename.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(file_storage, read_only=True, data_only=True)
            table = [list(r) for r in wb.active.iter_rows(values_only=True)]
        else:
            return [], "Unsupported file type. Please upload a .csv or .xlsx file."
    except Exception as exc:                                  # noqa: BLE001
        return [], f"Could not read that file: {str(exc)[:140]}"

    table = [r for r in table if r and any(str(c or "").strip() for c in r)]
    if not table:
        return [], "That file is empty."
    if len(table) - 1 > MAX_ROWS:
        return [], (f"That file has {len(table) - 1:,} rows — the limit is {MAX_ROWS:,}. "
                    "Please split it into smaller files.")

    cols = _header_map([str(c or "") for c in table[0]])
    if "name" not in cols or "date" not in cols:
        return [], ('The file must have a "Name" column and a "Date" column. '
                    "Download the template above to see the exact layout.")

    def cell(row, field):
        i = cols.get(field)
        if i is None or i >= len(row):
            return ""
        v = row[i]
        if isinstance(v, (date, datetime)):
            return v
        return str(v or "").strip()

    rows = []
    for n, raw in enumerate(table[1:], start=2):
        rows.append({
            "line": n,
            "name": str(cell(raw, "name") or "").strip(),
            "date": cell(raw, "date"),
            "end": cell(raw, "end"),
            "shift": str(cell(raw, "shift") or "").strip(),
            "leave": str(cell(raw, "leave") or "").strip(),
            "department": str(cell(raw, "department") or "").strip(),
            "section": str(cell(raw, "section") or "").strip(),
            "unit": str(cell(raw, "unit") or "").strip(),
            "note": str(cell(raw, "note") or "").strip()[:200],
        })
    return rows, None


# ------------------------------------------------------------------ validation
def build_preview(org_id: int, raw_rows: list[dict], *, place: dict,
                  mode: str | None, created_by=None) -> list[dict]:
    """Check every row and explain, in plain English, anything that is wrong.

    A row is rejected (never guessed at) when the person is unknown, the date
    cannot be read, the shift does not belong to this department's pattern, the
    same slot is already filled, or the person is already on leave that day.
    """
    people = staff_index(org_id)
    is_org = place.get("scope") == "ORG"
    allowed = [ORG_SHIFT] if is_org else shift_codes(mode)
    default_shift = allowed[0]

    existing = _existing_slots(org_id, place)
    org_existing_dates = {r.duty_date for r in
                          db.session.query(DutyRoster).filter_by(org_id=org_id).all()} \
        if place.get("scope") == "ORG" else set()
    leave_days = _existing_leave(org_id)
    seen: set[tuple] = set()
    out: list[dict] = []

    for row in raw_rows:
        errors: list[str] = []
        warnings: list[str] = []

        person = find_person(people, row["name"]) if row["name"] else None
        if not row["name"]:
            errors.append("No staff name on this line.")
        elif not person:
            errors.append(f"No active staff account matches “{row['name']}”. "
                          "Add them under Admin → Users first, or correct the spelling.")

        day = parse_date(row["date"])
        if not day:
            errors.append(f"Cannot read the date “{row['date']}”. Use 2026-09-14 or 14/09/2026.")

        end_day = parse_date(row["end"]) if row.get("end") else None
        if row.get("end") and not end_day:
            warnings.append(f"Ignoring unreadable end date “{row['end']}”.")
        if day and end_day and end_day < day:
            errors.append("The end date is before the start date.")
        span = ((end_day - day).days + 1) if (day and end_day) else 1
        if span > 120:
            errors.append(f"That is a {span}-day block — the limit is 120 days per line.")

        # leave or duty?
        leave_type = normalise_leave(row.get("leave") or "")
        if not leave_type:
            leave_type = normalise_leave(row.get("shift") or "")
        kind = "LEAVE" if leave_type else "DUTY"

        shift = (row.get("shift") or "").strip().upper()
        if kind == "DUTY" and is_org:
            # One Admin Manager per day, hospital-wide: any shift column is noise.
            shift = ORG_SHIFT
        elif kind == "DUTY":
            if not shift:
                shift = default_shift
                if len(allowed) > 1:
                    warnings.append(f"No shift given — assumed {default_shift}.")
            elif shift not in allowed:
                if shift in ALL_SHIFT_CODES:
                    errors.append(f"“{shift}” is not used by this roster pattern. "
                                  f"Allowed here: {', '.join(allowed)}.")
                else:
                    errors.append(f"“{shift}” is not a shift or a leave type. "
                                  f"Use one of {', '.join(allowed)}, or a leave type "
                                  f"({', '.join(LEAVE_CODES[:4])}…).")
        else:
            shift = "LEAVE"

        # per-day checks across the whole span
        if not errors and person and day:
            for d in days_between(day, end_day or day):
                if kind == "DUTY":
                    v = None if is_org else office_mode_violation(mode, d)
                    if v:
                        errors.append(v)
                        break
                    key = (d, person.id, shift)
                    if key in seen:
                        errors.append(f"This line repeats {person.name} on {d} ({shift}) "
                                      "— already earlier in the same file.")
                        break
                    if place.get("scope") == "ORG":
                        if d in org_existing_dates:
                            errors.append(f"{d} already has an Admin Manager on duty. "
                                          "Remove that entry first if you want to change it.")
                            break
                    elif key in existing:
                        errors.append(f"{person.name} is already rostered for {shift} on {d} here.")
                        break
                    if (person.id, d) in leave_days:
                        errors.append(f"{person.name} is on {LEAVE_LABELS.get(leave_days[(person.id, d)], 'leave')} "
                                      f"on {d} — they cannot be on duty that day.")
                        break
                else:
                    key = (d, person.id, "LEAVE")
                    if key in seen or (person.id, d) in leave_days:
                        warnings.append(f"{person.name} is already recorded as on leave on {d} "
                                        "— that day will be skipped.")
                        continue
                seen.add((d, person.id, shift))

        out.append({
            "line": row["line"], "name": row["name"],
            "person_id": person.id if person else None,
            "person_name": person.name if person else row["name"],
            "date": day.isoformat() if day else str(row["date"]),
            "end": end_day.isoformat() if end_day else "",
            "days": span,
            "kind": kind, "shift": shift, "leave_type": leave_type,
            "label": LEAVE_LABELS.get(leave_type, shift) if kind == "LEAVE" else shift,
            "note": row.get("note", ""),
            "errors": errors, "warnings": warnings, "ok": not errors,
        })
    return out


def _existing_slots(org_id: int, place: dict) -> set[tuple]:
    q = db.session.query(RosterEntry).filter_by(org_id=org_id, kind="DUTY",
                                                scope=place.get("scope", "DEPARTMENT"))
    for col in ("department_id", "section_id", "unit_id"):
        q = q.filter(getattr(RosterEntry, col) == place.get(col))
    return {(r.duty_date, r.user_id, r.shift) for r in q.all()}


def _existing_leave(org_id: int) -> dict[tuple, str]:
    rows = db.session.query(RosterEntry).filter_by(org_id=org_id, kind="LEAVE").all()
    return {(r.user_id, r.duty_date): (r.leave_type or "OFF") for r in rows}


# ------------------------------------------------------------------ preview store
def save_preview(org_id: int, rows: list[dict], meta: dict) -> str:
    from . import storage
    token = secrets.token_urlsafe(12)
    payload = json.dumps({"org_id": org_id, "at": now_naive().isoformat(),
                          "meta": meta, "rows": rows})
    storage.put(f"imports/roster-{token}.json", payload.encode("utf-8"),
                org_id=org_id, content_type="application/json")
    db.session.commit()
    return token


def load_preview(org_id: int, token: str):
    from . import storage
    if not token or "/" in token or ".." in token:
        return None, None
    data = storage.get(f"imports/roster-{token}.json")
    if not data:
        return None, None
    try:
        payload = json.loads(data.decode("utf-8"))
    except Exception:                                          # noqa: BLE001
        return None, None
    if payload.get("org_id") != org_id:                        # tenant isolation
        return None, None
    return payload.get("rows"), payload.get("meta") or {}


def discard_preview(org_id: int, token: str) -> None:
    from . import storage
    if token and "/" not in token and ".." not in token:
        storage.delete(f"imports/roster-{token}.json")
        db.session.commit()


# ------------------------------------------------------------------ commit
def commit_rows(org_id: int, rows: list[dict], *, place: dict, created_by_id: int) -> dict:
    """Write the approved rows. Re-checks everything — a preview can go stale."""
    added = skipped = 0
    leave_days = _existing_leave(org_id)
    existing = _existing_slots(org_id, place)
    org_dates = {r.duty_date for r in db.session.query(DutyRoster).filter_by(org_id=org_id).all()}

    for row in rows:
        if not row.get("ok") or not row.get("person_id"):
            skipped += 1
            continue
        start = parse_date(row["date"])
        end = parse_date(row.get("end")) or start
        if not start:
            skipped += 1
            continue
        uid = row["person_id"]
        for d in days_between(start, end):
            if row["kind"] == "LEAVE":
                if (uid, d) in leave_days:
                    skipped += 1
                    continue
                db.session.add(RosterEntry(
                    org_id=org_id, duty_date=d, user_id=uid, kind="LEAVE",
                    shift="LEAVE", leave_type=row.get("leave_type") or "OFF",
                    scope=place["scope"], department_id=place.get("department_id"),
                    section_id=place.get("section_id"), unit_id=place.get("unit_id"),
                    note=row.get("note") or None, source="import", created_by=created_by_id))
                leave_days[(uid, d)] = row.get("leave_type") or "OFF"
                added += 1
                continue

            if (uid, d) in leave_days:
                skipped += 1
                continue
            if place["scope"] == "ORG":
                if d in org_dates:
                    skipped += 1
                    continue
                db.session.add(DutyRoster(org_id=org_id, duty_date=d, user_id=uid,
                                          source="import", note=row.get("note") or None,
                                          created_by=created_by_id))
                org_dates.add(d)
                added += 1
                continue
            key = (d, uid, row["shift"])
            if key in existing:
                skipped += 1
                continue
            db.session.add(RosterEntry(
                org_id=org_id, duty_date=d, user_id=uid, kind="DUTY", shift=row["shift"],
                scope=place["scope"], department_id=place.get("department_id"),
                section_id=place.get("section_id"), unit_id=place.get("unit_id"),
                note=row.get("note") or None, source="import", created_by=created_by_id))
            existing.add(key)
            added += 1
    return {"added": added, "skipped": skipped}


# ------------------------------------------------------------------ reading
def load_roster(org_id: int, start: date, end: date, *, place: dict | None = None,
                include_org: bool = True) -> list[dict]:
    """Everything on the roster between two dates, as one sorted list."""
    out: list[dict] = []
    q = (db.session.query(RosterEntry)
         .filter(RosterEntry.org_id == org_id,
                 RosterEntry.duty_date >= start, RosterEntry.duty_date <= end))
    if place and place.get("scope") and place["scope"] != "ORG":
        q = q.filter(RosterEntry.scope == place["scope"])
        for col in ("department_id", "section_id", "unit_id"):
            if place.get(col):
                q = q.filter(getattr(RosterEntry, col) == place[col])
    elif place and place.get("scope") == "ORG":
        q = q.filter(RosterEntry.scope == "ORG")

    for r in q.all():
        out.append({"id": r.id, "kind": r.kind, "date": r.duty_date, "shift": r.shift,
                    "label": r.display_shift, "person": r.user.name if r.user else "—",
                    "place": r.place_label, "note": r.note or "", "source": r.source,
                    "table": "roster_entry", "leave_type": r.leave_type})

    if include_org and (not place or place.get("scope") in (None, "", "ORG")):
        for r in (db.session.query(DutyRoster)
                  .filter(DutyRoster.org_id == org_id, DutyRoster.duty_date >= start,
                          DutyRoster.duty_date <= end).all()):
            out.append({"id": r.id, "kind": "DUTY", "date": r.duty_date, "shift": "ADMIN",
                        "label": "Admin Manager on duty",
                        "person": r.user.name if r.user else "—",
                        "place": "Hospital-wide (Admin Manager)", "note": r.note or "",
                        "source": r.source or "manual", "table": "duty_roster",
                        "leave_type": None})

    out.sort(key=lambda x: (x["date"], x["kind"] == "LEAVE", x["label"], x["person"]))
    return out


def on_leave_between(org_id: int, start: date, end: date) -> list[RosterEntry]:
    return (db.session.query(RosterEntry)
            .filter(RosterEntry.org_id == org_id, RosterEntry.kind == "LEAVE",
                    RosterEntry.duty_date >= start, RosterEntry.duty_date <= end)
            .order_by(RosterEntry.duty_date).all())


# ------------------------------------------------------------------ templates
def template_csv(mode: str = "two_12h") -> str:
    """A starter file matching this department's own working pattern."""
    codes = shift_codes(mode)
    today = now_naive().date()
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Name", "Date", "End Date", "Shift", "Leave Type", "Section", "Unit", "Note"])
    w.writerow(["MRS ABATAN L.F", today.isoformat(), "", codes[0], "", "", "", ""])
    if len(codes) > 1:
        w.writerow(["CNO OGUNLEYE", today.isoformat(), "", codes[1], "", "", "", ""])
    w.writerow(["PHARM KAREEM", (today + timedelta(days=1)).isoformat(), "", codes[0], "",
                "", "", "Covering for Mrs Oba"])
    w.writerow(["MISS ADESANYA", (today + timedelta(days=2)).isoformat(),
                (today + timedelta(days=8)).isoformat(), "", "ANNUAL", "", "",
                "7 days annual leave"])
    w.writerow(["MR AFOLABI", (today + timedelta(days=3)).isoformat(), "", "", "SICK", "", "", ""])
    return buf.getvalue()


NOMINAL_HINT = (
    "Nominal roll files (Name / Department / Phone) create STAFF ACCOUNTS, not "
    "roster lines — use Admin → Users → Bulk upload for those. This page places "
    "people who already have accounts onto duty."
)


# ------------------------------------------------------------------ legacy migration
def migrate_legacy_entries(app) -> int:
    """Copy old two-column department roster rows into the unified roster, once.

    Runs at boot, is idempotent, and never deletes the original rows: if this
    ever went wrong the old table is still there, untouched.
    """
    moved = 0
    try:
        legacy = db.session.query(DeptRosterEntry).all()
    except Exception:                                          # noqa: BLE001 (table may not exist yet)
        return 0
    if not legacy:
        return 0
    have = {(r.org_id, r.duty_date, r.user_id, r.shift, r.department_id)
            for r in db.session.query(RosterEntry).filter_by(source="legacy").all()}
    for old in legacy:
        for uid in (old.staff1_user_id, old.staff2_user_id):
            if not uid:
                continue
            key = (old.org_id, old.duty_date, uid, old.shift, old.department_id)
            if key in have:
                continue
            db.session.add(RosterEntry(
                org_id=old.org_id, duty_date=old.duty_date, user_id=uid, kind="DUTY",
                shift=old.shift, scope="DEPARTMENT", department_id=old.department_id,
                source="legacy", created_by=old.created_by, created_at=old.created_at))
            have.add(key)
            moved += 1
    if moved:
        db.session.commit()
        app.logger.info("roster: migrated %s legacy department roster rows", moved)
    return moved
