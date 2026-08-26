"""Bulk staff upload — nominal roll, departmental list, or unit/section list.

WHY THIS EXISTS
---------------
Entering 200 staff by hand is 200 forms. The founder photographed a real duty
roster (Name / Department / Phone) — that is exactly the shape of data hospitals
already keep in Excel, so we accept it directly.

DESIGN NOTES
------------
* PREVIEW BEFORE WRITE. Nothing touches the database until a human has seen
  exactly what will happen to every row. Bad rows are shown with the reason and
  skipped, never guessed at.

* PREVIEW IS NOT STORED IN THE SESSION COOKIE. The existing roster import does
  that, and a Flask session cookie holds roughly 4KB — about 24 staff rows.
  A real nominal roll silently overflowed it. Previews here live in durable
  storage keyed by a short token.

* USERNAMES ARE GENERATED AND DE-DUPLICATED. Hospitals supply names, not
  usernames. "MRS ODEBE IDEHAI" becomes "odebe.idehai", and a clash becomes
  "odebe.idehai2".

* PASSWORDS ARE RANDOM AND MUST BE CHANGED. Nobody gets a predictable password,
  and every imported account starts with must_change_password=True.

* IMPORTED ACCOUNTS START UNAPPROVED. A bulk file should not silently grant
  hospital-wide access; an administrator approves them (Admin -> Users).
"""
from __future__ import annotations

import csv
import io
import json
import re
import secrets
import unicodedata

from .models import Department, User, db, now_naive

MAX_ROWS = 2000                 # a hospital nominal roll, with headroom

# Column aliases: hospitals label spreadsheets differently. Accept them all.
_ALIASES = {
    "name": ["name", "full name", "fullname", "staff name", "names", "staff",
             "employee name", "officer", "admin manager"],
    "department": ["department", "dept", "unit", "section", "department/unit",
                   "posting", "duty post", "assigned department"],
    "phone": ["phone", "phone no", "phone number", "mobile", "gsm", "telephone",
              "tel", "contact", "phone_no", "msisdn"],
    "email": ["email", "e-mail", "email address", "mail"],
    "role": ["role", "designation", "cadre", "position", "rank", "title"],
    "username": ["username", "user name", "login", "user id", "staff id"],
}

# Free-text designations mapped to system roles. Anything unrecognised becomes
# HOD (department-level access) rather than accidentally granting admin rights.
# ORDER MATTERS: the most specific titles are checked first. "Deputy Medical
# Director" must never be read as "Medical Director" — that would silently give
# a deputy the MD/CEO's authority. Bare "md"/"ceo" are matched as whole words
# only, for the same reason.
_ROLE_WORDS = [
    (("super admin", "system admin", "administrator", "sysadmin"), "SUPER_ADMIN"),
    (("dmd", "deputy medical director", "deputy md", "deputy director"), "DMD"),
    (("dcst", "director of clinical", "clinical services and training"), "DCST"),
    (("apex", "head of nursing", "chief nursing", "cno", "adns"), "APEX_NURSE"),
    (("head of admin", "admin & hr", "admin and hr", "hr manager",
      "human resources"), "HEAD_ADMIN_HR"),
    (("medical director", "chief executive", "chief medical director",
      "md", "ceo", "cmd"), "MD_CEO"),
    (("admin manager", "administrative manager", "duty manager"), "ADMIN_MANAGER"),
    (("hod", "head of department", "head of unit", "consultant", "doctor", "dr",
      "pharm", "nurse", "officer", "staff"), "HOD"),
]

_PHONE_OK = re.compile(r"^\+?\d{7,15}$")

# Real abbreviations from real Nigerian hospital duty rosters. Taken from an
# actual nominal roll: MEDICAL, PUB AFF OFF, ADMIN/HR, FIN/ACCTS, NUTRIT&DIET,
# INT AUDIT, HIMS... Without these, most rows import with no department.
_DEPT_ALIASES = {
    "medical": "Internal Medicine",
    "medicine": "Internal Medicine",
    "med": "Internal Medicine",
    "pub aff off": "Public Affairs",
    "pub aff": "Public Affairs",
    "public affairs office": "Public Affairs",
    "pro": "Public Affairs",
    "admin/hr": "Administration & Human Resources",
    "admin hr": "Administration & Human Resources",
    "admin": "Administration & Human Resources",
    "hr": "Administration & Human Resources",
    "fin/accts": "Finance & Accounts",
    "fin/acct": "Finance & Accounts",
    "fin accts": "Finance & Accounts",
    "finance": "Finance & Accounts",
    "accounts": "Finance & Accounts",
    "accts": "Finance & Accounts",
    "lab": "Laboratory",
    "nursing": "Nursing Services",
    "nurse": "Nursing Services",
    "hims": "Health Information Management (HIMS)",
    "him": "Health Information Management (HIMS)",
    "records": "Health Information Management (HIMS)",
    "medical records": "Health Information Management (HIMS)",
    "nutrit&diet": "Nutrition & Dietetics",
    "nutrit & diet": "Nutrition & Dietetics",
    "nutrition": "Nutrition & Dietetics",
    "diet": "Nutrition & Dietetics",
    "dietetics": "Nutrition & Dietetics",
    "int audit": "Internal Audit",
    "audit": "Internal Audit",
    "planning": "Planning, Research & Statistics",
    "stats": "Planning, Research & Statistics",
    "statistics": "Planning, Research & Statistics",
    "environmental": "Environmental Health",
    "environment": "Environmental Health",
    "sanitation": "Environmental Health",
    "engineering": "Engineering & Maintenance",
    "maintenance": "Engineering & Maintenance",
    "works": "Engineering & Maintenance",
    "catering": "Catering Services",
    "kitchen": "Catering Services",
    "dental": "Dental Services",
    "eye": "Ophthalmology (Eye Clinic)",
    "ophthalmology": "Ophthalmology (Eye Clinic)",
    "a&e": "Accident & Emergency",
    "a & e": "Accident & Emergency",
    "emergency": "Accident & Emergency",
    "casualty": "Accident & Emergency",
    "o&g": "Obstetrics & Gynaecology",
    "o & g": "Obstetrics & Gynaecology",
    "obs&gyn": "Obstetrics & Gynaecology",
    "maternity": "Obstetrics & Gynaecology",
    "anc": "Obstetrics & Gynaecology",
    "paeds": "Paediatrics",
    "peads": "Paediatrics",
    "children": "Paediatrics",
    "gopd": "Family Medicine / General Outpatient",
    "opd": "Family Medicine / General Outpatient",
    "outpatient": "Family Medicine / General Outpatient",
    "family medicine": "Family Medicine / General Outpatient",
    "x-ray": "Radiology / Imaging",
    "xray": "Radiology / Imaging",
    "radiology": "Radiology / Imaging",
    "imaging": "Radiology / Imaging",
    "scan": "Radiology / Imaging",
    "physio": "Physiotherapy",
    "theatre": "Surgery",
    "surgical": "Surgery",
    "security": "Security",
    "laundry": "Laundry",
    "mortuary": "Mortuary",
    "ict": "ICT",
    "it": "ICT",
    "pharm": "Pharmacy",
    "ent": "ENT (Ear, Nose & Throat)",
    "ortho": "Orthopaedics",
    "psychiatry": "Mental Health",
    "mental": "Mental Health",
}


def _alias_target(raw: str) -> str | None:
    """Resolve a hospital's shorthand to a standard department name."""
    key = re.sub(r"\s+", " ", (raw or "").strip().lower())
    if key in _DEPT_ALIASES:
        return _DEPT_ALIASES[key]
    squashed = key.replace(" ", "")
    for alias, target in _DEPT_ALIASES.items():
        if alias.replace(" ", "") == squashed:
            return target
    return None


# ------------------------------------------------------------------ helpers
def _norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9 ]", "", (h or "").strip().lower()).strip()


def _map_columns(headers: list[str]) -> dict[str, int]:
    """Map our field names to column positions, however the sheet is labelled."""
    found: dict[str, int] = {}
    cleaned = [_norm_header(h) for h in headers]
    for field, names in _ALIASES.items():
        for i, h in enumerate(cleaned):
            if h in names:
                found[field] = i
                break
    return found


def clean_phone(raw: str) -> str:
    return re.sub(r"[^\d+]", "", (raw or "").strip())


_SHORT_TITLES = {"md", "ceo", "cmd", "dmd", "dcst", "cno", "adns", "hod", "hr"}


def guess_role(raw: str) -> str:
    """Map a free-text designation to a system role.

    Unrecognised titles become HOD (department level), never an admin role:
    a spreadsheet must not be able to grant hospital-wide authority by accident.
    """
    low = " ".join((raw or "").strip().lower().split())
    if not low:
        return "HOD"
    words = set(re.findall(r"[a-z]+", low))
    for phrases, role in _ROLE_WORDS:
        for w in phrases:
            if w in _SHORT_TITLES:
                if w in words:            # whole word only: "md" not inside "medical"
                    return role
            elif w in low:
                return role
    return "HOD"


def make_username(full_name: str, taken: set[str]) -> str:
    """'MRS ODEBE IDEHAI' -> 'odebe.idehai'. Titles stripped, clashes numbered."""
    txt = unicodedata.normalize("NFKD", full_name or "")
    txt = txt.encode("ascii", "ignore").decode().lower()
    txt = re.sub(r"[^a-z ]", " ", txt)
    parts = [p for p in txt.split() if p]
    titles = {"mr", "mrs", "miss", "ms", "dr", "prof", "pharm", "engr", "arc",
              "chief", "alhaji", "alhaja", "rev", "pastor", "sir", "madam",
              "cno", "adns", "mallam"}
    parts = [p for p in parts if p not in titles] or parts or ["staff"]
    base = ".".join(parts[:2])[:40].strip(".") or "staff"
    candidate, n = base, 1
    while candidate in taken:
        n += 1
        candidate = f"{base}{n}"[:48]
    taken.add(candidate)
    return candidate


def new_password() -> str:
    """Random, readable-once temporary password. Always force-changed at login."""
    return secrets.token_urlsafe(9) + "A1!"


# ------------------------------------------------------------------ parsing
def parse_file(file_storage) -> tuple[list[dict], str | None]:
    """Read CSV/XLSX into raw dict rows. Returns (rows, error)."""
    name = (file_storage.filename or "").lower()
    rows: list[dict] = []
    try:
        if name.endswith(".csv"):
            text = file_storage.read().decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text))
            table = [r for r in reader if any((c or "").strip() for c in r)]
        elif name.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(file_storage, read_only=True, data_only=True)
            table = []
            for r in wb.active.iter_rows(values_only=True):
                vals = ["" if c is None else str(c).strip() for c in r]
                if any(vals):
                    table.append(vals)
        else:
            return [], "Please upload a .csv or .xlsx file."
    except Exception as exc:                             # noqa: BLE001
        return [], f"That file could not be read ({exc}). Please check the format."

    if not table:
        return [], "That file is empty."
    if len(table) > MAX_ROWS + 1:
        return [], (f"That file has more than {MAX_ROWS} rows. "
                    "Please split it into smaller files.")

    cols = _map_columns(table[0])
    if "name" not in cols:
        return [], ("Could not find a 'Name' column. The file needs a header row "
                    "with at least: Name, Department, Phone.")

    def cell(row, field):
        i = cols.get(field)
        if i is None or i >= len(row):
            return ""
        return str(row[i] or "").strip()

    for line_no, row in enumerate(table[1:], start=2):
        if not any((c or "").strip() for c in row):
            continue
        rows.append({
            "line": line_no,
            "name": cell(row, "name"),
            "department": cell(row, "department"),
            "phone": cell(row, "phone"),
            "email": cell(row, "email"),
            "role_raw": cell(row, "role"),
            "username": cell(row, "username"),
        })
    return rows, None


# ------------------------------------------------------------------ validation
def build_preview(org_id: int, raw_rows: list[dict], *,
                  default_department_id: int | None = None) -> list[dict]:
    """Validate every row and describe exactly what will happen to it."""
    depts = {d.name.strip().lower(): d for d in
             db.session.query(Department).filter_by(org_id=org_id).all()}
    existing_users = {u.username.lower() for u in db.session.query(User).all()}
    existing_phones = {(u.phone or "").strip(): u.name
                       for u in db.session.query(User).filter_by(org_id=org_id).all()
                       if u.phone}
    taken = set(existing_users)
    seen_names: set[str] = set()
    seen_phones: set[str] = set()

    default_dept = (db.session.get(Department, default_department_id)
                    if default_department_id else None)

    out = []
    for r in raw_rows:
        errors, warnings = [], []
        name = " ".join((r["name"] or "").split())
        if len(name) < 2:
            errors.append("Missing or too-short name")

        # department: named in the file, else the one chosen on the form
        dept = None
        dept_raw = (r["department"] or "").strip()
        if dept_raw:
            dept = depts.get(dept_raw.strip().lower())
            if dept is None:
                # 1) known hospital shorthand ("FIN/ACCTS" -> Finance & Accounts)
                target = _alias_target(dept_raw)
                if target and target.lower() in depts:
                    dept = depts[target.lower()]
                    warnings.append(f"'{dept_raw}' matched to '{dept.name}'")
            if dept is None:
                # 2) prefix / containment, longest match wins so "MED" does not
                #    beat a better candidate
                probe = dept_raw.strip().lower()[:6]
                best = None
                for key, d in depts.items():
                    if probe and (key.startswith(probe) or probe in key):
                        if best is None or len(key) < len(best[0]):
                            best = (key, d)
                if best:
                    dept = best[1]
                    warnings.append(f"'{dept_raw}' matched to '{dept.name}'")
            if dept is None:
                warnings.append(f"Unknown department '{dept_raw}' — will be left unassigned")
        elif default_dept is not None:
            dept = default_dept

        raw_phone = (r["phone"] or "").strip()
        phone = clean_phone(raw_phone)
        # A value like "not-a-number" cleans down to "", which used to slip
        # through silently — the admin never learned the number was discarded.
        if raw_phone and not phone:
            warnings.append(f"Phone '{raw_phone}' is not a valid number — left blank")
        if phone and not _PHONE_OK.match(phone):
            warnings.append(f"Phone '{raw_phone}' is not a valid number — left blank")
            phone = ""
        if phone and phone in seen_phones:
            warnings.append("Duplicate phone in this file")
        if phone and phone in existing_phones:
            warnings.append(f"Phone already used by {existing_phones[phone]}")

        key = name.strip().lower()
        if key and key in seen_names:
            errors.append("Duplicate name in this file")

        role = guess_role(r["role_raw"])
        username = (r["username"] or "").strip().lower() or make_username(name, taken)
        if username in existing_users:
            errors.append(f"Username '{username}' already exists")

        if not errors:
            seen_names.add(key)
            if phone:
                seen_phones.add(phone)

        out.append({
            "line": r["line"], "name": name, "username": username,
            "department_id": dept.id if dept else None,
            "department": dept.name if dept else "",
            "phone": phone, "email": (r["email"] or "").strip(),
            "role": role, "role_raw": r["role_raw"],
            "errors": errors, "warnings": warnings, "ok": not errors,
        })
    return out


# ------------------------------------------------------------------ preview store
def save_preview(org_id: int, rows: list[dict]) -> str:
    """Store a preview durably and return its token.

    Deliberately NOT the session cookie: a Flask session holds about 4KB, which
    is roughly 24 staff rows. A real nominal roll silently overflowed it.
    """
    from . import storage
    token = secrets.token_urlsafe(12)
    payload = json.dumps({"org_id": org_id, "at": now_naive().isoformat(), "rows": rows})
    storage.put(f"imports/{token}.json", payload.encode("utf-8"),
                org_id=org_id, content_type="application/json")
    db.session.commit()
    return token


def load_preview(org_id: int, token: str) -> list[dict] | None:
    from . import storage
    if not token or "/" in token or ".." in token:
        return None
    data = storage.get(f"imports/{token}.json")
    if not data:
        return None
    try:
        payload = json.loads(data.decode("utf-8"))
    except Exception:                                    # noqa: BLE001
        return None
    if payload.get("org_id") != org_id:                  # tenant isolation
        return None
    return payload.get("rows")


def discard_preview(org_id: int, token: str) -> None:
    from . import storage
    if token and "/" not in token:
        storage.delete(f"imports/{token}.json")
        db.session.commit()


# ------------------------------------------------------------------ commit
def commit_preview(org_id: int, rows: list[dict], *, created_by_id: int) -> dict:
    """Create the users. Returns a summary plus the temporary credentials."""
    created, skipped = [], 0
    taken = {u.username.lower() for u in db.session.query(User).all()}

    for row in rows:
        if not row.get("ok"):
            skipped += 1
            continue
        username = (row["username"] or "").lower()
        if not username or username in taken:
            skipped += 1
            continue
        pwd = new_password()
        u = User(
            org_id=org_id,
            username=username,
            name=row["name"],
            role=row["role"] if row["role"] in
            ("SUPER_ADMIN", "MD_CEO", "DMD", "DCST", "APEX_NURSE",
             "HEAD_ADMIN_HR", "ADMIN_MANAGER", "HOD") else "HOD",
            email=row.get("email") or None,
            phone=row.get("phone") or None,
            department_id=row.get("department_id"),
            approved=False,                 # an administrator approves them
            email_verified=False,           # they must prove they own the mailbox
            profile_completed=False,        # they fill their own staff card
            must_change_password=True,      # nobody keeps a generated password
        )
        u.set_password(pwd)
        db.session.add(u)
        taken.add(username)
        created.append({"name": row["name"], "username": username,
                        "password": pwd, "role": row["role"],
                        "department": row.get("department") or ""})
    db.session.flush()
    return {"created": created, "created_count": len(created), "skipped": skipped}


# ------------------------------------------------------------------ template
def template_csv(kind: str = "nominal") -> str:
    """A starter file the hospital can fill in, matching their own paperwork."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Name", "Department", "Phone", "Email", "Designation"])
    if kind == "department":
        w.writerow(["Dr. Tunde Bakare", "Internal Medicine", "08031234567",
                    "tunde@example.com", "HOD"])
        w.writerow(["Nurse Grace Obi", "Internal Medicine", "08039876543", "", "Nurse"])
    elif kind == "unit":
        w.writerow(["Mrs. Sadiq M.O", "Administration & Human Resources", "08084105130",
                    "", "Admin Officer"])
        w.writerow(["Mr. Afolabi", "Finance & Accounts", "08033901140", "", "Accountant"])
    else:
        w.writerow(["DR ADENIYI", "Internal Medicine", "08065226200", "", "Doctor"])
        w.writerow(["MRS ODEBE IDEHAI", "Public Affairs", "08028327098", "", "Officer"])
        w.writerow(["MISS SADIQ M.O", "Administration & Human Resources", "08084105130",
                    "", "Admin"])
        w.writerow(["CNO OGUNLEYE", "Nursing Services", "08062801586", "", "CNO"])
        w.writerow(["PHARM KAREEM", "Pharmacy", "09031737994", "", "Pharmacist"])
    return buf.getvalue()
