"""Duty roster module: calendar, manual entry, Excel/CSV import with validation."""
from __future__ import annotations

import csv
import io
from datetime import date, timedelta

from flask import (Blueprint, abort, flash, redirect, render_template, request,
                   session, url_for)
from flask_login import current_user

from ..audit import audit
from ..models import (DEPT_SHIFTS, Department, DeptRosterEntry, DutyRoster, User,
                      db, now_naive)
from ..security import require_role

bp = Blueprint("roster", __name__)


# Hard cap on imported rows. A 512 MB free instance can be knocked over by a
# spreadsheet with a million rows; refuse politely instead of dying.
MAX_IMPORT_ROWS = 2000

DATE_FORMATS = ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y")


def _parse_date(raw: str) -> date | None:
    raw = (raw or "").strip()
    for fmt in DATE_FORMATS:
        try:
            from datetime import datetime
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _org_admin_managers(org_id: int) -> list[User]:
    return (db.session.query(User)
            .filter(User.org_id == org_id, User.role.in_(("ADMIN_MANAGER", "SUPER_ADMIN")),
                    User.active.is_(True)).order_by(User.name).all())


# ------------------------------------------------------------------ view
@bp.get("/roster")
@require_role("ADMIN_MANAGER", "SUPER_ADMIN", "MD_CEO", "HOD")
def roster_view():
    today = now_naive().date()
    month_offset = request.args.get("m", type=int) or 0
    base = today.replace(day=1) + timedelta(days=32 * month_offset)
    start = base.replace(day=1)
    end = (start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    rows = (db.session.query(DutyRoster)
            .filter(DutyRoster.org_id == current_user.org_id,
                    DutyRoster.duty_date >= start, DutyRoster.duty_date <= end)
            .order_by(DutyRoster.duty_date).all())
    by_date = {r.duty_date: r for r in rows}
    days = []
    d = start
    while d <= end:
        days.append({"date": d, "entry": by_date.get(d), "today": d == today})
        d += timedelta(days=1)
    return render_template("roster.html", days=days, month_label=start.strftime("%B %Y"),
                           month_offset=month_offset,
                           admins=_org_admin_managers(current_user.org_id))


# ------------------------------------------------------------------ manual entry
@bp.post("/roster/manual")
@require_role("SUPER_ADMIN", "ADMIN_MANAGER")
def roster_manual():
    if not current_user.is_super:
        abort(403)  # roster management is an admin function
    d = _parse_date(request.form.get("date", ""))
    user_id = request.form.get("user_id", type=int)
    note = (request.form.get("note") or "").strip()
    user = db.session.get(User, user_id) if user_id else None
    if not d:
        flash("Please provide a valid duty date.", "error")
        return redirect(url_for("roster.roster_view"))
    if not user or user.org_id != current_user.org_id or user.role != "ADMIN_MANAGER":
        flash("Please select a valid Admin Manager.", "error")
        return redirect(url_for("roster.roster_view"))
    existing = db.session.query(DutyRoster).filter_by(org_id=current_user.org_id, duty_date=d).first()
    if existing:
        flash(f"A roster entry already exists for {d} ({existing.user.name}). "
              "Edit or delete it first — duplicates are not allowed.", "error")
        return redirect(url_for("roster.roster_view"))
    db.session.add(DutyRoster(org_id=current_user.org_id, duty_date=d, user_id=user.id,
                              source="manual", note=note, created_by=current_user.id))
    audit("ROSTER_MANUAL_ADD", "roster", None, {"date": str(d), "user": user.name})
    db.session.commit()
    flash(f"Roster entry added for {d} — {user.name}.", "success")
    return redirect(url_for("roster.roster_view"))


@bp.post("/roster/<int:rid>/delete")
@require_role("SUPER_ADMIN")
def roster_delete(rid: int):
    r = db.session.get(DutyRoster, rid)
    if not r or r.org_id != current_user.org_id:
        abort(404)
    audit("ROSTER_DELETED", "roster", r.id, {"date": str(r.duty_date), "user": r.user.name})
    db.session.delete(r)
    db.session.commit()
    flash("Roster entry deleted.", "success")
    return redirect(url_for("roster.roster_view"))


@bp.post("/roster/<int:rid>/reassign")
@require_role("SUPER_ADMIN")
def roster_reassign(rid: int):
    r = db.session.get(DutyRoster, rid)
    if not r or r.org_id != current_user.org_id:
        abort(404)
    user = db.session.get(User, request.form.get("user_id", type=int) or 0)
    if not user or user.role != "ADMIN_MANAGER" or user.org_id != current_user.org_id:
        flash("Select a valid Admin Manager.", "error")
        return redirect(url_for("roster.roster_view"))
    old = r.user.name
    r.user_id = user.id
    audit("ROSTER_REASSIGNED", "roster", r.id, {"date": str(r.duty_date), "old": old, "new": user.name})
    db.session.commit()
    flash(f"Duty on {r.duty_date} reassigned to {user.name}.", "success")
    return redirect(url_for("roster.roster_view"))


# ------------------------------------------------------------------ import
@bp.get("/roster/import")
@require_role("SUPER_ADMIN")
def roster_import_form():
    return render_template("roster_import.html")


@bp.post("/roster/import")
@require_role("SUPER_ADMIN")
def roster_import_parse():
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Please choose an Excel (.xlsx) or CSV file.", "error")
        return redirect(url_for("roster.roster_import_form"))
    name = file.filename.lower()
    rows_raw: list[tuple[str, str]] = []
    try:
        if name.endswith(".csv"):
            text = file.read().decode("utf-8-sig", errors="replace")
            for row in csv.DictReader(io.StringIO(text)):
                keys = {(k or "").strip().lower(): v for k, v in row.items()}
                nm = keys.get("name") or keys.get("admin manager") or keys.get("admin_manager") or ""
                dt = keys.get("date") or keys.get("duty date") or keys.get("duty_date") or ""
                note = keys.get("note") or keys.get("duty assignment") or ""
                rows_raw.append(((nm or "").strip(), (dt or "").strip(), (note or "").strip()))
                if len(rows_raw) > MAX_IMPORT_ROWS:
                    flash(f"That file has too many rows (limit {MAX_IMPORT_ROWS}). "
                          "Please split it into smaller files.", "error")
                    return redirect(url_for("roster.roster_import_form"))
        elif name.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]

            def col(*candidates):
                for cand in candidates:
                    if cand in header:
                        return header.index(cand)
                return None

            i_name = col("name", "admin manager", "admin_manager")
            i_date = col("date", "duty date", "duty_date")
            i_note = col("note", "duty assignment", "assignment")
            if i_name is None or i_date is None:
                flash('The spreadsheet must contain "Name" and "Date" columns.', "error")
                return redirect(url_for("roster.roster_import_form"))
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None:
                    continue
                nm = str(row[i_name] or "").strip()
                dt = str(row[i_date] or "").strip()
                if hasattr(row[i_date], "date"):
                    dt = row[i_date].date().isoformat()
                note = str(row[i_note] or "").strip() if i_note is not None and i_note < len(row) else ""
                rows_raw.append((nm, dt, note))
                if len(rows_raw) > MAX_IMPORT_ROWS:
                    flash(f"That file has too many rows (limit {MAX_IMPORT_ROWS}). "
                          "Please split it into smaller files.", "error")
                    return redirect(url_for("roster.roster_import_form"))
        else:
            flash("Unsupported file type. Please upload .xlsx or .csv.", "error")
            return redirect(url_for("roster.roster_import_form"))
    except Exception as exc:  # noqa: BLE001
        flash(f"Could not read the file: {str(exc)[:120]}", "error")
        return redirect(url_for("roster.roster_import_form"))

    # ------- validation
    admins = {u.name.lower(): u for u in _org_admin_managers(current_user.org_id)}
    existing_dates = {r.duty_date for r in db.session.query(DutyRoster).filter_by(org_id=current_user.org_id).all()}
    seen_in_file: set[date] = set()
    preview, error_count = [], 0
    for idx, (nm, dt, note) in enumerate(rows_raw, start=2):
        errs = []
        parsed_date = None
        user = None
        if not nm:
            errs.append("Missing Admin Manager name")
        else:
            user = admins.get(nm.lower())
            if not user:
                errs.append(f"Unknown Admin Manager: {nm}")
        if not dt:
            errs.append("Missing date")
        else:
            parsed_date = _parse_date(dt)
            if not parsed_date:
                errs.append(f"Invalid date: {dt}")
            else:
                if parsed_date in seen_in_file:
                    errs.append(f"Duplicate date in file: {parsed_date}")
                if parsed_date in existing_dates:
                    errs.append(f"Date already rostered: {parsed_date}")
        if errs:
            error_count += 1
        else:
            seen_in_file.add(parsed_date)
        preview.append({"row": idx, "name": nm, "date": dt, "note": note, "errors": errs,
                        "ok": not errs})
    session["roster_import_preview"] = preview
    session["roster_import_errors"] = error_count
    return render_template("roster_import_preview.html", preview=preview, error_count=error_count,
                           valid_count=len(preview) - error_count)


@bp.post("/roster/import/confirm")
@require_role("SUPER_ADMIN")
def roster_import_confirm():
    preview = session.get("roster_import_preview")
    if not preview:
        flash("Import preview expired. Please upload the file again.", "error")
        return redirect(url_for("roster.roster_import_form"))
    admins = {u.name.lower(): u for u in _org_admin_managers(current_user.org_id)}
    added, skipped = 0, 0
    existing_dates = {r.duty_date for r in db.session.query(DutyRoster).filter_by(org_id=current_user.org_id).all()}
    for row in preview:
        if row["errors"]:
            skipped += 1
            continue
        d = _parse_date(row["date"])
        user = admins.get(row["name"].lower())
        if not d or not user or d in existing_dates:
            skipped += 1
            continue
        db.session.add(DutyRoster(org_id=current_user.org_id, duty_date=d, user_id=user.id,
                                  source="import", note=row.get("note") or None,
                                  created_by=current_user.id))
        existing_dates.add(d)
        added += 1
    audit("ROSTER_IMPORTED", "roster", None, {"added": added, "skipped": skipped})
    db.session.commit()
    session.pop("roster_import_preview", None)
    flash(f"Import complete: {added} entries added, {skipped} skipped.", "success")
    return redirect(url_for("roster.roster_view"))


# ================================================================ DEPARTMENT ROSTERS (§upgrade)
def _can_manage_dept(user, dept) -> bool:
    return user.is_super or (user.is_hod and dept.hod_user_id == user.id)


def _parse_day(raw: str):
    return _parse_date(raw)


@bp.get("/dept-roster/template")
@require_role("SUPER_ADMIN", "HOD", "ADMIN_MANAGER", "MD_CEO")
def dept_roster_template():
    mode = request.args.get("mode", "two_12h")
    shifts = [s[0] for s in DEPT_SHIFTS.get(mode, DEPT_SHIFTS["two_12h"])]
    lines = ["Date,Shift,Staff1,Staff2"]
    lines.append(f"2026-01-01,{shifts[0]},Full Name One,Full Name Two")
    if len(shifts) > 1:
        lines.append(f"2026-01-01,{shifts[1]},Full Name Three,")
    from flask import Response
    return Response("\n".join(lines) + "\n", mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=dept-roster-template-{mode}.csv"})


@bp.get("/dept-roster")
@require_role("SUPER_ADMIN", "HOD", "ADMIN_MANAGER", "MD_CEO")
def dept_roster():
    from ..models import DeptRosterEntry, User
    depts = db.session.query(Department).filter_by(org_id=current_user.org_id, active=True)
    if current_user.is_hod:
        depts = depts.filter(Department.hod_user_id == current_user.id)
    depts = depts.order_by(Department.name).all()
    dept_id = request.args.get("dept", type=int)
    dept = None
    if dept_id:
        dept = db.session.get(Department, dept_id)
        if not dept or dept.org_id != current_user.org_id or not _can_manage_dept(current_user, dept):
            dept = depts[0] if depts else None
    elif depts:
        dept = depts[0]
    month_offset = request.args.get("m", type=int) or 0
    entries, month_label = [], now_naive().date().strftime("%B %Y")
    if dept:
        today = now_naive().date()
        start = (today.replace(day=1) + timedelta(days=32 * month_offset)).replace(day=1)
        end = (start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
        month_label = start.strftime("%B %Y")
        entries = (db.session.query(DeptRosterEntry)
                   .filter(DeptRosterEntry.department_id == dept.id,
                           DeptRosterEntry.duty_date >= start, DeptRosterEntry.duty_date <= end)
                   .order_by(DeptRosterEntry.duty_date, DeptRosterEntry.shift).all())
    staff = (db.session.query(User).filter_by(org_id=current_user.org_id, active=True)
             .order_by(User.name).all())
    return render_template("dept_roster.html", depts=depts, dept=dept, entries=entries,
                           staff=staff, month_offset=month_offset, month_label=month_label,
                           shifts=DEPT_SHIFTS,
                           can_manage=bool(dept and _can_manage_dept(current_user, dept)))


def _validate_dept_entry(org_id, dept, raw_date, shift, name1, name2):
    """Returns (errors, day, staff1, staff2)."""
    from ..models import User
    errors = []
    day = _parse_day(raw_date)
    if not day:
        errors.append("Invalid or missing date.")
    allowed = [s[0] for s in DEPT_SHIFTS.get(dept.roster_mode or "two_12h", [])]
    if shift not in allowed:
        errors.append(f"Shift must be one of {', '.join(allowed)} for this department's roster system.")
    staff1 = db.session.query(User).filter_by(org_id=org_id, name=name1.strip(), active=True).first() if name1.strip() else None
    if not staff1:
        errors.append(f"Unknown staff: {name1 or '(missing)'}")
    staff2 = None
    if (dept.roster_staff_per_shift or 1) >= 2 and name2.strip():
        staff2 = db.session.query(User).filter_by(org_id=org_id, name=name2.strip(), active=True).first()
        if not staff2:
            errors.append(f"Unknown staff: {name2}")
    elif name2.strip():
        errors.append("This department is configured for ONE staff on duty per shift.")
    if staff1 and staff2 and staff1.id == staff2.id:
        errors.append("Staff1 and Staff2 must be different people.")
    return errors, day, staff1, staff2


@bp.post("/dept-roster/add")
@require_role("SUPER_ADMIN", "HOD")
def dept_roster_add():
    from ..models import DeptRosterEntry
    dept = db.session.get(Department, request.form.get("department_id", type=int) or 0)
    if not dept or dept.org_id != current_user.org_id or not _can_manage_dept(current_user, dept):
        abort(403)
    errors, day, s1, s2 = _validate_dept_entry(
        current_user.org_id, dept, request.form.get("duty_date", ""),
        request.form.get("shift", ""), request.form.get("staff1", ""),
        request.form.get("staff2", ""))
    if not errors:
        dup = (db.session.query(DeptRosterEntry)
               .filter_by(department_id=dept.id, duty_date=day,
                          shift=request.form.get("shift")).first())
        if dup:
            errors.append("An entry for this date & shift already exists — edit it instead.")
    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("roster.dept_roster", dept=dept.id))
    db.session.add(DeptRosterEntry(org_id=current_user.org_id, department_id=dept.id,
                                   duty_date=day, shift=request.form.get("shift"),
                                   staff1_user_id=s1.id, staff2_user_id=s2.id if s2 else None,
                                   source="manual", created_by=current_user.id))
    audit("DEPT_ROSTER_ADD", "dept_roster", None,
          {"dept": dept.name, "date": str(day), "shift": request.form.get("shift")})
    db.session.commit()
    flash(f"Roster entry added for {dept.name} ({day}, {request.form.get('shift')}).", "success")
    return redirect(url_for("roster.dept_roster", dept=dept.id))


@bp.post("/dept-roster/<int:rid>/edit")
@require_role("SUPER_ADMIN", "HOD")
def dept_roster_edit(rid: int):
    from ..models import DeptRosterEntry
    r = db.session.get(DeptRosterEntry, rid)
    if not r or r.org_id != current_user.org_id or not _can_manage_dept(current_user, r.department):
        abort(403)
    errors, day, s1, s2 = _validate_dept_entry(
        current_user.org_id, r.department, request.form.get("duty_date", ""),
        request.form.get("shift", ""), request.form.get("staff1", ""),
        request.form.get("staff2", ""))
    if not errors:
        dup = (db.session.query(DeptRosterEntry)
               .filter(DeptRosterEntry.department_id == r.department_id,
                       DeptRosterEntry.duty_date == day,
                       DeptRosterEntry.shift == request.form.get("shift"),
                       DeptRosterEntry.id != rid).first())
        if dup:
            errors.append("Another entry already covers this date & shift.")
    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("roster.dept_roster", dept=r.department_id))
    old = (str(r.duty_date), r.shift, r.staff1.name, r.staff2.name if r.staff2 else None)
    r.duty_date = day
    r.shift = request.form.get("shift")
    r.staff1_user_id = s1.id
    r.staff2_user_id = s2.id if s2 else None
    audit("DEPT_ROSTER_EDIT", "dept_roster", rid, {"old": old,
          "new": (str(day), r.shift, s1.name, s2.name if s2 else None)})
    db.session.commit()
    flash("Roster entry updated.", "success")
    return redirect(url_for("roster.dept_roster", dept=r.department_id))


@bp.post("/dept-roster/<int:rid>/delete")
@require_role("SUPER_ADMIN", "HOD")
def dept_roster_delete(rid: int):
    from ..models import DeptRosterEntry
    r = db.session.get(DeptRosterEntry, rid)
    if not r or r.org_id != current_user.org_id or not _can_manage_dept(current_user, r.department):
        abort(403)
    audit("DEPT_ROSTER_DELETE", "dept_roster", rid,
          {"dept": r.department.name, "date": str(r.duty_date), "shift": r.shift})
    db.session.delete(r)
    db.session.commit()
    flash("Roster entry deleted.", "success")
    return redirect(url_for("roster.dept_roster", dept=r.department_id))


@bp.post("/dept-roster/import")
@require_role("SUPER_ADMIN", "HOD")
def dept_roster_import():
    """CSV/XLSX upload: Date,Shift,Staff1,Staff2 — validated with preview-less quick report."""
    import csv as _csv, io as _io
    dept = db.session.get(Department, request.form.get("department_id", type=int) or 0)
    if not dept or dept.org_id != current_user.org_id or not _can_manage_dept(current_user, dept):
        abort(403)
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Choose a CSV or XLSX file.", "error")
        return redirect(url_for("roster.dept_roster", dept=dept.id))
    rows = []
    name = file.filename.lower()
    try:
        if name.endswith(".csv"):
            text = file.read().decode("utf-8-sig", errors="replace")
            for row in _csv.DictReader(_io.StringIO(text)):
                k = {(x or "").strip().lower(): (v or "").strip() for x, v in row.items()}
                rows.append((k.get("date", ""), k.get("shift", ""), k.get("staff1", ""), k.get("staff2", "")))
                if len(rows) > MAX_IMPORT_ROWS:
                    flash(f"That file has too many rows (limit {MAX_IMPORT_ROWS}). "
                          "Please split it into smaller files.", "error")
                    return redirect(url_for("roster.dept_roster", dept=dept.id))
        elif name.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(file, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    d = row[0]
                    rows.append((d.date().isoformat() if hasattr(d, "date") else str(d),
                                 str(row[1] or ""), str(row[2] or ""), str(row[3] or "")))
        else:
            flash("Unsupported file type — use .csv or .xlsx.", "error")
            return redirect(url_for("roster.dept_roster", dept=dept.id))
    except Exception as exc:  # noqa: BLE001
        flash(f"Could not read file: {str(exc)[:120]}", "error")
        return redirect(url_for("roster.dept_roster", dept=dept.id))
    added, skipped, problems = 0, 0, 0
    from ..models import DeptRosterEntry
    for raw_date, shift, n1, n2 in rows:
        if shift.lower().startswith("date") or not raw_date:
            skipped += 1
            continue
        errors, day, s1, s2 = _validate_dept_entry(current_user.org_id, dept, raw_date,
                                                   shift.strip().upper(), n1, n2)
        if errors:
            problems += 1
            continue
        dup = db.session.query(DeptRosterEntry).filter_by(
            department_id=dept.id, duty_date=day, shift=shift.strip().upper()).first()
        if dup:
            problems += 1
            continue
        db.session.add(DeptRosterEntry(org_id=current_user.org_id, department_id=dept.id,
                                       duty_date=day, shift=shift.strip().upper(),
                                       staff1_user_id=s1.id, staff2_user_id=s2.id if s2 else None,
                                       source="import", created_by=current_user.id))
        added += 1
    audit("DEPT_ROSTER_IMPORTED", "dept_roster", None,
          {"dept": dept.name, "added": added, "problems": problems})
    db.session.commit()
    flash(f"Import finished for {dept.name}: {added} added, {problems} rejected, {skipped} skipped.",
          "success" if not problems else "info")
    return redirect(url_for("roster.dept_roster", dept=dept.id))
